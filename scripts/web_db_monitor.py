#!/usr/bin/env python3
"""
Веб-интерфейс для мониторинга базы данных MariaDB
"""

from flask import Flask, render_template_string, jsonify, request, send_file
import subprocess
import json
from datetime import datetime
import threading
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile

app = Flask(__name__)

# Конфигурация базы данных
DB_CONFIG = {
    'user': 'crewlife_user',
    'password': 'andrei8002012',
    'database': 'crewlife_prod'
}

def run_mysql_query(query):
    """Выполнение SQL запроса"""
    try:
        cmd = f"mysql -u {DB_CONFIG['user']} -p{DB_CONFIG['password']} {DB_CONFIG['database']} -e \"{query}\""
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
        if result.returncode == 0:
            return result.stdout.strip()
        else:
            return f"Ошибка: {result.stderr}"
    except Exception as e:
        return f"Ошибка выполнения: {e}"

def create_excel_report():
    """Создание Excel отчета с данными из базы"""
    try:
        # Создаем новую рабочую книгу
        wb = openpyxl.Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Лист "Статистика"
        stats_sheet = wb.create_sheet("Статистика")
        
        # Получаем статистику
        stats_data = get_database_stats_for_excel()
        
        stats_sheet['A1'] = "Параметр"
        stats_sheet['B1'] = "Значение"
        stats_sheet['C1'] = "Дата обновления"
        
        # Применяем стили к заголовкам
        for col in ['A1', 'B1', 'C1']:
            stats_sheet[col].font = header_font
            stats_sheet[col].fill = header_fill
            stats_sheet[col].alignment = header_alignment
            stats_sheet[col].border = thin_border
        
        # Заполняем данные статистики
        row = 2
        for key, value in stats_data.items():
            if key != 'timestamp':
                stats_sheet[f'A{row}'] = key.replace('_', ' ').title()
                stats_sheet[f'B{row}'] = value
                stats_sheet[f'C{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Применяем границы
                for col in ['A', 'B', 'C']:
                    stats_sheet[f'{col}{row}'].border = thin_border
                row += 1
        
        # Автоподбор ширины колонок
        for column in stats_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            stats_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 2. Лист "Пользователи"
        users_sheet = wb.create_sheet("Пользователи")
        
        # Заголовки для пользователей
        users_headers = ["ID", "Табельный номер", "ФИО", "Должность", "Локация", "Статус", "Дата создания", "Последний вход"]
        for col, header in enumerate(users_headers, 1):
            cell = users_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Получаем данные пользователей
        users_data = get_users_for_excel()
        for row, user in enumerate(users_data, 2):
            users_sheet.cell(row=row, column=1, value=user.get('id', ''))
            users_sheet.cell(row=row, column=2, value=user.get('employee_id', ''))
            users_sheet.cell(row=row, column=3, value=user.get('full_name', ''))
            users_sheet.cell(row=row, column=4, value=user.get('position', ''))
            users_sheet.cell(row=row, column=5, value=user.get('location', ''))
            users_sheet.cell(row=row, column=6, value='Активен' if user.get('is_active') else 'Неактивен')
            users_sheet.cell(row=row, column=7, value=user.get('created_at', ''))
            users_sheet.cell(row=row, column=8, value=user.get('last_login', ''))
            
            # Применяем границы
            for col in range(1, 9):
                users_sheet.cell(row=row, column=col).border = thin_border
        
        # Автоподбор ширины колонок для пользователей
        for column in users_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            users_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 3. Лист "Заявки"
        requests_sheet = wb.create_sheet("Заявки")
        
        # Заголовки для заявок
        requests_headers = ["ID", "Тип заявки", "Название", "Описание", "Приоритет", "Статус", "Дата создания", "Дата обновления"]
        for col, header in enumerate(requests_headers, 1):
            cell = requests_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Получаем данные заявок
        requests_data = get_requests_for_excel()
        for row, request in enumerate(requests_data, 2):
            requests_sheet.cell(row=row, column=1, value=request.get('id', ''))
            requests_sheet.cell(row=row, column=2, value=request.get('request_type', ''))
            requests_sheet.cell(row=row, column=3, value=request.get('title', ''))
            requests_sheet.cell(row=row, column=4, value=request.get('description', ''))
            requests_sheet.cell(row=row, column=5, value=request.get('priority', ''))
            requests_sheet.cell(row=row, column=6, value=request.get('status', ''))
            requests_sheet.cell(row=row, column=7, value=request.get('created_at', ''))
            requests_sheet.cell(row=row, column=8, value=request.get('updated_at', ''))
            
            # Применяем границы
            for col in range(1, 9):
                requests_sheet.cell(row=row, column=col).border = thin_border
        
        # Автоподбор ширины колонок для заявок
        for column in requests_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            requests_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # 4. Лист "Логи"
        logs_sheet = wb.create_sheet("Логи")
        
        # Заголовки для логов
        logs_headers = ["ID", "Действие", "Детали", "IP адрес", "User Agent", "Дата создания"]
        for col, header in enumerate(logs_headers, 1):
            cell = logs_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Получаем данные логов
        logs_data = get_logs_for_excel()
        for row, log in enumerate(logs_data, 2):
            logs_sheet.cell(row=row, column=1, value=log.get('id', ''))
            logs_sheet.cell(row=row, column=2, value=log.get('action', ''))
            logs_sheet.cell(row=row, column=3, value=log.get('details', ''))
            logs_sheet.cell(row=row, column=4, value=log.get('ip_address', ''))
            logs_sheet.cell(row=row, column=5, value=log.get('user_agent', ''))
            logs_sheet.cell(row=row, column=6, value=log.get('created_at', ''))
            
            # Применяем границы
            for col in range(1, 7):
                logs_sheet.cell(row=row, column=col).border = thin_border
        
        # Автоподбор ширины колонок для логов
        for column in logs_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            logs_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл во временную директорию
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"crewlife_report_{timestamp}.xlsx"
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        
        wb.save(temp_path)
        
        return temp_path, filename
        
    except Exception as e:
        print(f"Ошибка создания Excel отчета: {e}")
        return None, None

def get_database_stats_for_excel():
    """Получение статистики для Excel"""
    try:
        # Активные пользователи
        users_result = run_mysql_query("SELECT COUNT(*) FROM users WHERE is_active = 1")
        active_users = users_result.split('\n')[-1] if users_result else "0"
        
        # Всего заявок
        requests_result = run_mysql_query("SELECT COUNT(*) FROM requests")
        total_requests = requests_result.split('\n')[-1] if requests_result else "0"
        
        # Заявки в обработке
        pending_result = run_mysql_query("SELECT COUNT(*) FROM requests WHERE status = 'pending'")
        pending_requests = pending_result.split('\n')[-1] if pending_result else "0"
        
        # Размер базы данных
        size_result = run_mysql_query("SELECT ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS 'DB Size in MB' FROM information_schema.tables WHERE table_schema = 'crewlife_prod'")
        db_size_mb = size_result.split('\n')[-1] if size_result else "0"
        
        return {
            'active_users': active_users,
            'total_requests': total_requests,
            'pending_requests': pending_requests,
            'db_size_mb': db_size_mb,
            'timestamp': datetime.now().isoformat()
        }
    except Exception as e:
        return {'error': str(e)}

def get_users_for_excel():
    """Получение пользователей для Excel"""
    try:
        result = run_mysql_query("SELECT id, employee_id, full_name, position, location, is_active, created_at, last_login FROM users ORDER BY created_at DESC")
        if "Ошибка" in result:
            return []
        
        lines = result.split('\n')[1:]  # Пропускаем заголовок
        users = []
        for line in lines:
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 8:
                    users.append({
                        'id': parts[0],
                        'employee_id': parts[1],
                        'full_name': parts[2],
                        'position': parts[3],
                        'location': parts[4],
                        'is_active': int(parts[5]) == 1,
                        'created_at': parts[6],
                        'last_login': parts[7]
                    })
        return users
    except Exception as e:
        return []

def get_requests_for_excel():
    """Получение заявок для Excel"""
    try:
        result = run_mysql_query("SELECT id, request_type, title, description, priority, status, created_at, updated_at FROM requests ORDER BY created_at DESC")
        if "Ошибка" in result:
            return []
        
        lines = result.split('\n')[1:]  # Пропускаем заголовок
        requests = []
        for line in lines:
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 8:
                    requests.append({
                        'id': parts[0],
                        'request_type': parts[1],
                        'title': parts[2],
                        'description': parts[3],
                        'priority': parts[4],
                        'status': parts[5],
                        'created_at': parts[6],
                        'updated_at': parts[7]
                    })
        return requests
    except Exception as e:
        return []

def get_logs_for_excel():
    """Получение логов для Excel"""
    try:
        result = run_mysql_query("SELECT id, action, details, ip_address, user_agent, created_at FROM logs ORDER BY created_at DESC LIMIT 100")
        if "Ошибка" in result:
            return []
        
        lines = result.split('\n')[1:]  # Пропускаем заголовок
        logs = []
        for line in lines:
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 6:
                    logs.append({
                        'id': parts[0],
                        'action': parts[1],
                        'details': parts[2],
                        'ip_address': parts[3],
                        'user_agent': parts[4],
                        'created_at': parts[5]
                    })
        return logs
    except Exception as e:
        return []

# HTML шаблон
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CrewLife - Мониторинг БД</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container { 
            max-width: 1200px; 
            margin: 0 auto; 
            background: white; 
            border-radius: 15px; 
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            overflow: hidden;
        }
        .header { 
            background: linear-gradient(135deg, #2c3e50, #34495e); 
            color: white; 
            padding: 30px; 
            text-align: center; 
        }
        .header h1 { font-size: 2.5em; margin-bottom: 10px; }
        .header p { opacity: 0.8; font-size: 1.1em; }
        .stats-grid { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); 
            gap: 20px; 
            padding: 30px; 
        }
        .stat-card { 
            background: #f8f9fa; 
            padding: 25px; 
            border-radius: 10px; 
            text-align: center; 
            border-left: 5px solid #3498db;
            transition: transform 0.3s ease;
        }
        .stat-card:hover { transform: translateY(-5px); }
        .stat-card h3 { color: #2c3e50; margin-bottom: 10px; font-size: 1.2em; }
        .stat-card .value { font-size: 2.5em; font-weight: bold; color: #3498db; }
        .section { 
            margin: 30px; 
            background: #f8f9fa; 
            border-radius: 10px; 
            padding: 25px; 
        }
        .section h2 { color: #2c3e50; margin-bottom: 20px; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        .table-container { 
            overflow-x: auto; 
            background: white; 
            border-radius: 8px; 
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        table { 
            width: 100%; 
            border-collapse: collapse; 
        }
        th, td { 
            padding: 12px; 
            text-align: left; 
            border-bottom: 1px solid #ddd; 
        }
        th { 
            background: #3498db; 
            color: white; 
            font-weight: 600; 
        }
        tr:hover { background: #f5f5f5; }
        .refresh-btn { 
            background: #27ae60; 
            color: white; 
            border: none; 
            padding: 12px 25px; 
            border-radius: 5px; 
            cursor: pointer; 
            font-size: 1em; 
            margin: 10px 0;
            transition: background 0.3s ease;
        }
        .refresh-btn:hover { background: #229954; }
        .auto-refresh { 
            background: #e74c3c; 
            color: white; 
            border: none; 
            padding: 12px 25px; 
            border-radius: 5px; 
            cursor: pointer; 
            font-size: 1em; 
            margin: 10px 0;
            transition: background 0.3s ease;
        }
        .auto-refresh:hover { background: #c0392b; }
        .auto-refresh.active { background: #27ae60; }
        .status-indicator { 
            display: inline-block; 
            width: 12px; 
            height: 12px; 
            border-radius: 50%; 
            margin-right: 8px; 
        }
        .status-online { background: #27ae60; }
        .status-offline { background: #e74c3c; }
        .timestamp { 
            color: #7f8c8d; 
            font-size: 0.9em; 
            text-align: center; 
            margin: 20px 0; 
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🔍 Мониторинг базы данных</h1>
            <p>CrewLife Production Database Monitor</p>
        </div>
        
        <div class="stats-grid">
            <div class="stat-card">
                <h3>👥 Активные пользователи</h3>
                <div class="value" id="active-users">-</div>
            </div>
            <div class="stat-card">
                <h3>📝 Всего заявок</h3>
                <div class="value" id="total-requests">-</div>
            </div>
            <div class="stat-card">
                <h3>⏳ В обработке</h3>
                <div class="value" id="pending-requests">-</div>
            </div>
            <div class="stat-card">
                <h3>💾 Размер БД</h3>
                <div class="value" id="db-size">-</div>
            </div>
        </div>
        
        <div class="section">
            <h2>📊 Статистика в реальном времени</h2>
            <button class="refresh-btn" onclick="refreshData()">🔄 Обновить</button>
            <button class="auto-refresh" id="auto-refresh-btn" onclick="toggleAutoRefresh()">⏰ Автообновление</button>
            <button class="refresh-btn" onclick="exportToExcel()" style="background: #e67e22;">📊 Экспорт в Excel</button>
            <div class="timestamp" id="last-update">Последнее обновление: -</div>
        </div>
        
        <div class="section">
            <h2>👥 Пользователи</h2>
            <div class="table-container">
                <table id="users-table">
                    <thead>
                        <tr>
                            <th>ID</th>
                            <th>Табельный номер</th>
                            <th>ФИО</th>
                            <th>Должность</th>
                            <th>Локация</th>
                            <th>Статус</th>
                        </tr>
                    </thead>
                    <tbody id="users-tbody">
                        <tr><td colspan="6">Загрузка...</td></tr>
                    </tbody>
                </table>
            </div>
        </div>
        
        <div class="section">
            <h2>📝 Заявки</h2>
            <div class="table-container">
                <table id="requests-table">
                    <thead>
                        <tr>
                            <th>ID</th>
                            <th>Тип</th>
                            <th>Название</th>
                            <th>Приоритет</th>
                            <th>Статус</th>
                            <th>Создана</th>
                        </tr>
                    </thead>
                    <tbody id="requests-tbody">
                        <tr><td colspan="6">Загрузка...</td></tr>
                    </tbody>
                </table>
            </div>
        </div>
        
        <div class="section">
            <h2>📋 Последние действия</h2>
            <div class="table-container">
                <table id="logs-table">
                    <thead>
                        <tr>
                            <th>Действие</th>
                            <th>Детали</th>
                            <th>Время</th>
                        </tr>
                    </thead>
                    <tbody id="logs-tbody">
                        <tr><td colspan="3">Загрузка...</td></tr>
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script>
        let autoRefreshInterval = null;
        let isAutoRefresh = false;
        
        function refreshData() {
            fetch('/api/stats')
                .then(response => response.json())
                .then(data => {
                    document.getElementById('active-users').textContent = data.active_users;
                    document.getElementById('total-requests').textContent = data.total_requests;
                    document.getElementById('pending-requests').textContent = data.pending_requests;
                    document.getElementById('db-size').textContent = data.db_size_mb + ' MB';
                    document.getElementById('last-update').textContent = 'Последнее обновление: ' + new Date().toLocaleString('ru-RU');
                })
                .catch(error => console.error('Ошибка:', error));
            
            fetch('/api/users')
                .then(response => response.json())
                .then(data => {
                    const tbody = document.getElementById('users-tbody');
                    tbody.innerHTML = '';
                    data.forEach(user => {
                        const row = tbody.insertRow();
                        row.innerHTML = `
                            <td>${user.id}</td>
                            <td>${user.employee_id}</td>
                            <td>${user.full_name}</td>
                            <td>${user.position}</td>
                            <td>${user.location}</td>
                            <td><span class="status-indicator ${user.is_active ? 'status-online' : 'status-offline'}"></span>${user.is_active ? 'Активен' : 'Неактивен'}</td>
                        `;
                    });
                })
                .catch(error => console.error('Ошибка:', error));
            
            fetch('/api/requests')
                .then(response => response.json())
                .then(data => {
                    const tbody = document.getElementById('requests-tbody');
                    tbody.innerHTML = '';
                    if (data.length === 0) {
                        tbody.innerHTML = '<tr><td colspan="6">Нет заявок</td></tr>';
                    } else {
                        data.forEach(request => {
                            const row = tbody.insertRow();
                            row.innerHTML = `
                                <td>${request.id}</td>
                                <td>${request.request_type}</td>
                                <td>${request.title}</td>
                                <td>${request.priority}</td>
                                <td>${request.status}</td>
                                <td>${new Date(request.created_at).toLocaleString('ru-RU')}</td>
                            `;
                        });
                    }
                })
                .catch(error => console.error('Ошибка:', error));
            
            fetch('/api/logs')
                .then(response => response.json())
                .then(data => {
                    const tbody = document.getElementById('logs-tbody');
                    tbody.innerHTML = '';
                    if (data.length === 0) {
                        tbody.innerHTML = '<tr><td colspan="3">Нет логов</td></tr>';
                    } else {
                        data.forEach(log => {
                            const row = tbody.insertRow();
                            row.innerHTML = `
                                <td>${log.action}</td>
                                <td>${log.details}</td>
                                <td>${new Date(log.created_at).toLocaleString('ru-RU')}</td>
                            `;
                        });
                    }
                })
                .catch(error => console.error('Ошибка:', error));
        }
        
        function toggleAutoRefresh() {
            const btn = document.getElementById('auto-refresh-btn');
            if (isAutoRefresh) {
                clearInterval(autoRefreshInterval);
                btn.textContent = '⏰ Автообновление';
                btn.classList.remove('active');
                isAutoRefresh = false;
            } else {
                autoRefreshInterval = setInterval(refreshData, 5000);
                btn.textContent = '⏹️ Остановить';
                btn.classList.add('active');
                isAutoRefresh = true;
            }
        }
        
        function exportToExcel() {
            // Показываем индикатор загрузки
            const btn = event.target;
            const originalText = btn.textContent;
            btn.textContent = '⏳ Создание отчета...';
            btn.disabled = true;
            
            // Создаем ссылку для скачивания
            const link = document.createElement('a');
            link.href = '/api/export/excel';
            link.download = 'crewlife_report.xlsx';
            link.style.display = 'none';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            
            // Восстанавливаем кнопку
            setTimeout(() => {
                btn.textContent = originalText;
                btn.disabled = false;
            }, 2000);
        }
        
        // Загружаем данные при загрузке страницы
        document.addEventListener('DOMContentLoaded', refreshData);
    </script>
</body>
</html>
"""

@app.route('/')
def index():
    """Главная страница"""
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/stats')
def api_stats():
    """API для получения статистики"""
    try:
        # Активные пользователи
        users_result = run_mysql_query("SELECT COUNT(*) FROM users WHERE is_active = 1")
        active_users = users_result.split('\n')[-1] if users_result else "0"
        
        # Всего заявок
        requests_result = run_mysql_query("SELECT COUNT(*) FROM requests")
        total_requests = requests_result.split('\n')[-1] if requests_result else "0"
        
        # Заявки в обработке
        pending_result = run_mysql_query("SELECT COUNT(*) FROM requests WHERE status = 'pending'")
        pending_requests = pending_result.split('\n')[-1] if pending_result else "0"
        
        # Размер базы данных
        size_result = run_mysql_query("SELECT ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS 'DB Size in MB' FROM information_schema.tables WHERE table_schema = 'crewlife_prod'")
        db_size_mb = size_result.split('\n')[-1] if size_result else "0"
        
        return jsonify({
            'active_users': active_users,
            'total_requests': total_requests,
            'pending_requests': pending_requests,
            'db_size_mb': db_size_mb,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/users')
def api_users():
    """API для получения пользователей"""
    try:
        result = run_mysql_query("SELECT id, employee_id, full_name, position, location, is_active FROM users ORDER BY created_at DESC")
        if "Ошибка" in result:
            return jsonify([])
        
        lines = result.split('\n')[1:]  # Пропускаем заголовок
        users = []
        for line in lines:
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 6:
                    users.append({
                        'id': parts[0],
                        'employee_id': parts[1],
                        'full_name': parts[2],
                        'position': parts[3],
                        'location': parts[4],
                        'is_active': int(parts[5]) == 1
                    })
        return jsonify(users)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/requests')
def api_requests():
    """API для получения заявок"""
    try:
        result = run_mysql_query("SELECT id, request_type, title, priority, status, created_at FROM requests ORDER BY created_at DESC LIMIT 20")
        if "Ошибка" in result:
            return jsonify([])
        
        lines = result.split('\n')[1:]  # Пропускаем заголовок
        requests = []
        for line in lines:
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 6:
                    requests.append({
                        'id': parts[0],
                        'request_type': parts[1],
                        'title': parts[2],
                        'priority': parts[3],
                        'status': parts[4],
                        'created_at': parts[5]
                    })
        return jsonify(requests)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/logs')
def api_logs():
    """API для получения логов"""
    try:
        result = run_mysql_query("SELECT action, details, created_at FROM logs ORDER BY created_at DESC LIMIT 10")
        if "Ошибка" in result:
            return jsonify([])
        
        lines = result.split('\n')[1:]  # Пропускаем заголовок
        logs = []
        for line in lines:
            if line.strip():
                parts = line.split('\t')
                if len(parts) >= 3:
                    logs.append({
                        'action': parts[0],
                        'details': parts[1],
                        'created_at': parts[2]
                    })
        return jsonify(logs)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel')
def export_excel():
    """API для экспорта данных в Excel"""
    try:
        temp_path, filename = create_excel_report()
        if temp_path and filename:
            return send_file(
                temp_path,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({'error': 'Ошибка создания Excel отчета'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("🌐 Запуск веб-мониторинга базы данных...")
    print("📊 Доступ: http://localhost:5001")
    print("🛑 Нажмите Ctrl+C для остановки")
    app.run(host='0.0.0.0', port=5001, debug=False)
